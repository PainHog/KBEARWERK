"""Excel helpers built on openpyxl.

Design goals:
* Never lose her data. We open the real workbook, append/patch, and save; we
  never rewrite it from scratch.
* Be forgiving about headers. Column order and exact capitalisation vary between
  her sheets, so we match headers case-insensitively and ignore surrounding
  whitespace.
* Give a *clear* error when the file is open in Excel (the #1 real-world
  failure), instead of a cryptic PermissionError.
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except Exception:  # pragma: no cover - only if dependency missing
    _OPENPYXL_OK = False


class ExcelError(Exception):
    """A user-friendly Excel problem (message is safe to show in the UI)."""


class ExcelBusyError(ExcelError):
    """The file is open in Excel / locked. She just needs to close it and retry."""


class ExcelUnavailableError(ExcelError):
    """The file/folder can't be reached (e.g. cloud server desync).

    Callers should preserve her work locally instead of losing it.
    """


def _require_openpyxl() -> None:
    if not _OPENPYXL_OK:
        raise ExcelError(
            "The Excel engine (openpyxl) is not installed.\n"
            "Reinstall the app, or run: pip install openpyxl"
        )


def _norm(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _parent_available(path: str) -> bool:
    parent = os.path.dirname(os.path.abspath(path)) or "."
    return os.path.isdir(parent)


def _load(path: str):
    _require_openpyxl()
    if not os.path.exists(path):
        # Distinguish "the whole folder is gone" (server desync) from "just this
        # file is missing" so callers can protect her work in the first case.
        if not _parent_available(path):
            raise ExcelUnavailableError(
                f"Can't reach this location right now (the cloud folder may be "
                f"disconnected):\n{path}"
            )
        raise ExcelError(f"Spreadsheet not found:\n{path}")
    try:
        return load_workbook(path)
    except PermissionError:
        raise ExcelBusyError(
            f"Can't open the spreadsheet because it looks like it's already "
            f"open in Excel:\n{os.path.basename(path)}\n\n"
            f"Please close it in Excel and try again."
        )
    except OSError as exc:
        raise ExcelUnavailableError(
            f"Couldn't reach the spreadsheet (the cloud location may be "
            f"disconnected):\n{os.path.basename(path)}\n\n({exc})"
        )
    except Exception as exc:  # pragma: no cover - defensive
        raise ExcelError(f"Couldn't read the spreadsheet:\n{exc}")


def _save(wb, path: str) -> None:
    try:
        wb.save(path)
    except PermissionError:
        raise ExcelBusyError(
            f"Can't save the spreadsheet because it's open in Excel:\n"
            f"{os.path.basename(path)}\n\nPlease close it in Excel and try again."
        )
    except OSError as exc:
        raise ExcelUnavailableError(
            f"Couldn't save to the cloud location (it may be disconnected):\n"
            f"{os.path.basename(path)}\n\n({exc})"
        )


def _pick_sheet(wb, sheet_name: Optional[str]):
    if sheet_name:
        for name in wb.sheetnames:
            if _norm(name) == _norm(sheet_name):
                return wb[name]
        raise ExcelError(f"The tab '{sheet_name}' was not found in the workbook.")
    return wb.active


def _header_map(ws) -> Dict[str, int]:
    """Map normalised header text -> 1-based column index, from row 1."""
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        key = _norm(val)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def ensure_workbook(path: str, headers: List[str], sheet_name: Optional[str] = None) -> None:
    """Create the workbook with a header row if it does not exist yet."""
    _require_openpyxl()
    if os.path.exists(path):
        return
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    if sheet_name:
        ws.title = sheet_name
    for i, head in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=head)
    _save(wb, path)


def read_headers(path: str, sheet_name: Optional[str] = None) -> List[str]:
    wb = _load(path)
    ws = _pick_sheet(wb, sheet_name)
    return [
        str(ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]


def read_rows(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """Return every data row as a dict keyed by the (original) header text."""
    wb = _load(path)
    ws = _pick_sheet(wb, sheet_name)
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue  # skip blank rows
        rows.append({headers[i]: values[i] for i in range(len(headers))})
    return rows


def append_row(
    path: str,
    values: Dict[str, Any],
    sheet_name: Optional[str] = None,
    unique_key: Optional[str] = None,
) -> str:
    """Append a row, matching ``values`` keys to existing headers by name.

    ``values`` is ``{header_text: cell_value}``. Headers are matched
    case-insensitively; keys with no matching column are ignored (so callers can
    pass a superset of fields and each sheet takes what it has).

    If ``unique_key`` is given and a row already has that column equal to the
    value in ``values``, that row is **updated** instead of a new one appended
    (used to avoid duplicate proposals/projects). Returns "added" or "updated".
    """
    wb = _load(path)
    ws = _pick_sheet(wb, sheet_name)
    header_map = _header_map(ws)
    if not header_map:
        raise ExcelError(
            f"The sheet has no header row, so the app doesn't know which column "
            f"is which:\n{os.path.basename(path)}"
        )

    # Try to update an existing row first.
    if unique_key and _norm(unique_key) in header_map and _norm(unique_key) in {_norm(k) for k in values}:
        key_col = header_map[_norm(unique_key)]
        target_val = None
        for k, v in values.items():
            if _norm(k) == _norm(unique_key):
                target_val = v
                break
        for r in range(2, ws.max_row + 1):
            if _norm(ws.cell(row=r, column=key_col).value) == _norm(target_val):
                _write_row(ws, r, values, header_map)
                _save(wb, path)
                return "updated"

    # Otherwise append at the first truly empty row.
    new_row = _first_empty_row(ws)
    _write_row(ws, new_row, values, header_map)
    _save(wb, path)
    return "added"


def _first_empty_row(ws) -> int:
    r = ws.max_row
    # max_row can over-report if there are stray formats; walk back over blanks.
    while r >= 2:
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            break
        r -= 1
    return r + 1


def _write_row(ws, row: int, values: Dict[str, Any], header_map: Dict[str, int]) -> None:
    for key, val in values.items():
        col = header_map.get(_norm(key))
        if col:
            ws.cell(row=row, column=col, value=val)
